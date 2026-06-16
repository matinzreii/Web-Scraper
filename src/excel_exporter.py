"""
Excel export module with professional formatting.

This module exports cleaned data to Excel files with:
- Auto-adjusted column widths
- Color-coded headers
- Number formatting
- Filter-enabled tables
"""

from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import OUTPUT_DIR, OUTPUT_FILE_PREFIX
from src.logger_config import setup_logger

logger = setup_logger(__name__)


class ExcelExporter:
    """
    Export pandas DataFrame to professionally formatted Excel files.
    """

    def __init__(self, df: pd.DataFrame, base_filename: str = None):
        """
        Initialize the exporter.

        Args:
            df: DataFrame to export.
            base_filename: Base name for the output file (without extension).
        """
        self.df = df
        self.base_filename = base_filename or OUTPUT_FILE_PREFIX
        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.output_path = OUTPUT_DIR / f"{self.base_filename}_{self.timestamp}.xlsx"

    def export(self) -> Path:
        """
        Export DataFrame to Excel with professional formatting.

        Returns:
            Path to the created Excel file.
        """
        if self.df.empty:
            logger.warning("No data to export")
            return None

        logger.info(f"Exporting {len(self.df)} records to Excel")

        # Create Excel writer
        with pd.ExcelWriter(self.output_path, engine="openpyxl") as writer:
            # Write main data sheet
            self.df.to_excel(writer, sheet_name="Competitor Data", index=False)

        # Apply formatting
        self._apply_formatting()

        logger.info(f"Excel file saved: {self.output_path}")
        return self.output_path

    def _apply_formatting(self):
        """
        Apply professional formatting to the Excel file.
        """
        workbook = load_workbook(self.output_path)
        worksheet = workbook["Competitor Data"]

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # Format header row
        for col_idx, cell in enumerate(worksheet[1], start=1):
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Apply number formatting to numeric columns
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.border = thin_border

                # Format price_numeric as currency
                if cell.column_letter == self._get_column_letter(worksheet, "price_numeric"):
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.00'
                        cell.font = Font(color="006100")

                # Format rating_normalized as number with 1 decimal
                if cell.column_letter == self._get_column_letter(worksheet, "rating_normalized"):
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '0.0'

        # Add a summary sheet
        self._add_summary_sheet(workbook)

        workbook.save(self.output_path)

    def _get_column_letter(self, worksheet, column_name: str) -> str:
        """
        Get the column letter for a given column name.

        Args:
            worksheet: Worksheet object.
            column_name: Name of the column.

        Returns:
            Column letter (e.g., 'A', 'B').
        """
        for col_idx, cell in enumerate(worksheet[1], start=1):
            if cell.value == column_name:
                return get_column_letter(col_idx)
        return None

    def _add_summary_sheet(self, workbook):
        """
        Add a summary sheet with statistics about the scraped data.
        """
        summary_data = {
            "Metric": [
                "Total Products Scraped",
                "Unique Competitors",
                "Products with Valid Prices",
                "Average Price (numeric only)",
                "Average Rating",
                "In-Stock Products",
                "Out-of-Stock Products",
                "Scrape Timestamp",
            ],
            "Value": [
                len(self.df),
                self.df["competitor_name"].nunique() if "competitor_name" in self.df.columns else "N/A",
                self.df["price_numeric"].notna().sum() if "price_numeric" in self.df.columns else "N/A",
                f"${self.df['price_numeric'].mean():.2f}" if "price_numeric" in self.df.columns and self.df['price_numeric'].notna().any() else "N/A",
                self.df["rating_normalized"].mean() if "rating_normalized" in self.df.columns and self.df['rating_normalized'].notna().any() else "N/A",
                self.df["availability_standardized"].eq("In Stock").sum() if "availability_standardized" in self.df.columns else "N/A",
                self.df["availability_standardized"].eq("Out of Stock").sum() if "availability_standardized" in self.df.columns else "N/A",
                self.timestamp,
            ],
        }

        summary_df = pd.DataFrame(summary_data)

        with pd.ExcelWriter(self.output_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            summary_df.to_excel(writer, sheet_name="Summary", index=False)

        # Format summary sheet
        summary_ws = workbook["Summary"]
        summary_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        for cell in summary_ws[1]:
            cell.fill = summary_header_fill
            cell.font = Font(bold=True, color="FFFFFF")

        for col in summary_ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            summary_ws.column_dimensions[col_letter].width = min(max_length + 2, 30)